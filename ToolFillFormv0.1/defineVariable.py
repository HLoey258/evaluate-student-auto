# Reading Excel
EXCEL_FILE = "formb4b9.xlsx"
TEACHER_DETAIL = "Thông tin GV"
DETAIL = "Nội dung nhận xét"

from openpyxl import load_workbook

book = load_workbook(EXCEL_FILE)

sheet = book[DETAIL]
teacher_sheet = book[TEACHER_DETAIL]

# Link
test = "https://google.com"
formb4b9 = "https://docs.google.com/forms/d/e/1FAIpQLSdhmCXo_17SP0oDj82XmEvp8RFJyofVL8J5Iqss73EZojTdXQ/viewform"
tekyForm = "https://erp.teky.edu.vn/web/login"


# Constant - Teacher Detail
TEACHER_NAME = teacher_sheet["A2"].value
TEACHER_CODE = teacher_sheet["B2"].value
CENTER_CODE = teacher_sheet["D2"].value
# Day Report
DAY_REPORT = teacher_sheet["F2"].value
DAY_START = teacher_sheet["E2"].value


# Report code
LESSON = teacher_sheet["C2"].value


def getReportCode(lesson):
    match lesson:
        case 4:
            return 3
        case 9:
            return 4
        case 12:
            return 5
        case "Upsale":
            return 6
        case "upsale":
            return 6
        case default:
            return 3


REPORT_CODE = getReportCode(LESSON)

print("Ngày bắt đàu học phần: ", DAY_START)
print("Ngày BCCK: ", DAY_REPORT)
print("Báo cáo buổi số : ", LESSON)

# Text Field
XPATH_TeacherName = "/html/body/div/div[2]/form/div[2]/div/div[2]/div[1]/div/div/div[2]/div/div[1]/div/div[1]/input"
XPATH_TeacherCode = "/html/body/div/div[2]/form/div[2]/div/div[2]/div[2]/div/div/div[2]/div/div[1]/div/div[1]/input"
XPATH_DROPDOWN_CENTER_CODE = "/html/body/div/div[2]/form/div[2]/div/div[2]/div[3]/div/div/div[2]/div/div[1]/div[1]/div[1]"
XPATH_DROPDOWN_REPORT_CODE = "/html/body/div/div[2]/form/div[2]/div/div[2]/div[8]/div/div/div[2]/div/div[1]/div[1]/div[1]"
XPATH_DATE_START = "/html/body/div/div[2]/form/div[2]/div/div[2]/div[6]/div/div/div[2]/div/div/div[2]/div[1]/div/div[1]/input"
XPATH_DATE_REPORT = "/html/body/div/div[2]/form/div[2]/div/div[2]/div[7]/div/div/div[2]/div/div/div[2]/div[1]/div/div[1]/input"


# Next Button
XPATH_NEXTBUTTON = "/html/body/div/div[2]/form/div[2]/div/div[3]/div[1]/div[1]/div/span"

# Submit Button
XPATH_SUBMITBUTTON = (
    "/html/body/div/div[2]/form/div[2]/div/div[3]/div[1]/div[1]/div[2]/span"
)

# Append data to raw_data
raw_data = []
for row in sheet.iter_rows(min_row=2, max_col=6, values_only=True):
    raw_data.append(row)
